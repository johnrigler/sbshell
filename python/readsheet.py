from openpyxl import load_workbook
wb = load_workbook('doc2.xlsx')
print wb.get_sheet_names
for sheet in wb:
     print(sheet.title)
 
ws = wb["Sheet1"]
print(ws)
print(ws['A1'].value)
sheet_ranges = wb['range names']
print(sheet_ranges['A10'].value)
