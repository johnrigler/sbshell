import openpyxl

avg = [10, 20, 25, 5, 32, 7]
avg2 = []

wb = openpyxl.load_workbook('/home/jrigler/windblows/empty.xlsx')
ws = wb.active

for row, entry in enumerate(avg, start=2):
    ws.cell(row=row, column=1, value=entry)
    avg2.append(entry)

print(ws['A2'].value)
print(ws.cell(row=3,column=1).value)
    
print avg2
wb.save('/home/jrigler/windblows/out.xlsx')
