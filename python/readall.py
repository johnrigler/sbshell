from openpyxl import Workbook, load_workbook

wb=load_workbook("/home/jrigler/windblows/doc2.xlsx", read_only=True)
sheet_ranges=wb['PrivAlert Terms']

for row in sheet_ranges.iter_rows(row_offset=1): 
    for cell in row:
        print(cell.value)
